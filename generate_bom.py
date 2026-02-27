"""
Script pour générer un BOM (Bill of Materials) en croisant les données
de l'arborescence (tree_data.xlsx) avec les autres fichiers Excel.
"""
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Chemins des fichiers
BASE_PATH = r"c:\Users\Administrateur\Desktop\PFE"

# Fichiers sources
TREE_DATA_FILE = os.path.join(BASE_PATH, "tree_data.xlsx")
PMP_2022_FILE = os.path.join(BASE_PATH, "PMP 2022 (1).xls")
PMP_PLF_FILE = os.path.join(BASE_PATH, "PMP Condit PLF 2021 (1).xlsx")
CONSOMATION_FILE = os.path.join(BASE_PATH, "CL_EXPORT_ETAT_CONSOMATION3060118986151578542 (1).xls")
LISTE_CONTROLE_FILE = os.path.join(BASE_PATH, "LISTE DE CONTRÔLE DISPOSITIF DE SECURITE (1).xlsx")

# Fichier de sortie BOM
BOM_OUTPUT_FILE = os.path.join(BASE_PATH, "bom_output.xlsx")

def analyze_all_files():
    """Analyse la structure de tous les fichiers Excel"""
    results = []
    
    files = {
        "tree_data.xlsx": TREE_DATA_FILE,
        "PMP 2022": PMP_2022_FILE,
        "PMP PLF 2021": PMP_PLF_FILE,
        "Consommation": CONSOMATION_FILE,
        "Liste Controle": LISTE_CONTROLE_FILE
    }
    
    for name, path in files.items():
        print(f"\n{'='*60}")
        print(f"FICHIER: {name}")
        print(f"Chemin: {path}")
        print(f"Existe: {os.path.exists(path)}")
        print('='*60)
        
        if os.path.exists(path):
            try:
                xl = pd.ExcelFile(path)
                print(f"Feuilles: {xl.sheet_names}")
                
                for sheet in xl.sheet_names[:2]:
                    df = pd.read_excel(path, sheet_name=sheet)
                    print(f"\n--- Feuille: {sheet} ---")
                    print(f"Shape: {df.shape}")
                    print(f"Colonnes: {list(df.columns)}")
                    print(f"\nTypes:")
                    print(df.dtypes)
                    print(f"\nAperçu (3 premières lignes):")
                    print(df.head(3).to_string())
                    
                    # Identifier les colonnes potentielles de jointure (Part Number, Code, etc.)
                    potential_keys = [col for col in df.columns if any(kw in str(col).lower() for kw in ['part', 'code', 'ref', 'article', 'num', 'id', 'position'])]
                    if potential_keys:
                        print(f"\nColonnes potentielles de jointure: {potential_keys}")
                        
            except Exception as e:
                print(f"ERREUR: {e}")
        else:
            print("FICHIER NON TROUVÉ")
    
    return results

def load_tree_data():
    """Charge les données de l'arborescence depuis tree_data.xlsx"""
    if not os.path.exists(TREE_DATA_FILE):
        print(f"ERREUR: {TREE_DATA_FILE} n'existe pas")
        return None
    
    df = pd.read_excel(TREE_DATA_FILE)
    print(f"\nArborescence chargée: {df.shape[0]} lignes, {df.shape[1]} colonnes")
    print(f"Colonnes: {list(df.columns)}")
    return df

def load_source_files():
    """Charge tous les fichiers sources pour le croisement"""
    sources = {}
    
    # PMP 2022
    if os.path.exists(PMP_2022_FILE):
        try:
            sources['pmp_2022'] = pd.read_excel(PMP_2022_FILE)
            print(f"PMP 2022 chargé: {sources['pmp_2022'].shape}")
        except Exception as e:
            print(f"Erreur PMP 2022: {e}")
    
    # PMP PLF 2021
    if os.path.exists(PMP_PLF_FILE):
        try:
            sources['pmp_plf'] = pd.read_excel(PMP_PLF_FILE)
            print(f"PMP PLF chargé: {sources['pmp_plf'].shape}")
        except Exception as e:
            print(f"Erreur PMP PLF: {e}")
    
    # Consommation
    if os.path.exists(CONSOMATION_FILE):
        try:
            sources['consommation'] = pd.read_excel(CONSOMATION_FILE)
            print(f"Consommation chargé: {sources['consommation'].shape}")
        except Exception as e:
            print(f"Erreur Consommation: {e}")
    
    # Liste Contrôle
    if os.path.exists(LISTE_CONTROLE_FILE):
        try:
            sources['liste_controle'] = pd.read_excel(LISTE_CONTROLE_FILE)
            print(f"Liste Contrôle chargé: {sources['liste_controle'].shape}")
        except Exception as e:
            print(f"Erreur Liste Contrôle: {e}")
    
    return sources

def generate_bom(tree_df, sources):
    """
    Génère le BOM en croisant l'arborescence avec les sources.
    """
    if tree_df is None:
        return None
    
    bom_df = tree_df.copy()
    
    # TODO: Implémenter la logique de croisement basée sur l'analyse des fichiers
    # Pour l'instant, on retourne l'arborescence telle quelle
    
    return bom_df

def save_bom_with_formatting(bom_df, output_path):
    """
    Sauvegarde le BOM avec le même design que tree_data.xlsx
    """
    # Sauvegarder d'abord sans mise en forme
    bom_df.to_excel(output_path, index=False)
    
    # Appliquer la mise en forme
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Style pour l'en-tête
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Bordures
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Appliquer le style à l'en-tête
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Appliquer les bordures à toutes les cellules
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"\nBOM sauvegardé dans: {output_path}")

if __name__ == "__main__":
    print("="*60)
    print("GÉNÉRATION DU BOM - ANALYSE DES FICHIERS")
    print("="*60)
    
    # Étape 1: Analyser tous les fichiers
    analyze_all_files()
    
    # Étape 2: Charger l'arborescence
    tree_df = load_tree_data()
    
    # Étape 3: Charger les sources
    sources = load_source_files()
    
    # Étape 4: Générer le BOM
    if tree_df is not None:
        bom_df = generate_bom(tree_df, sources)
        if bom_df is not None:
            save_bom_with_formatting(bom_df, BOM_OUTPUT_FILE)
    
    print("\n" + "="*60)
    print("FIN DU SCRIPT")
    print("="*60)
