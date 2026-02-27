import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import pandas as pd
import uuid
import os

# Nom du fichier Excel pour la sauvegarde
EXCEL_FILE = "tree_data_internal copy.xlsx"

class TreeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestionnaire d'Arbre - Part Number & Description")
        self.root.geometry("900x600")

        # Cadre pour les boutons
        button_frame = tk.Frame(self.root)
        button_frame.pack(fill=tk.X, padx=10, pady=5)

        # Boutons d'action
        tk.Button(button_frame, text="Ajouter Racine", command=self.add_root).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Ajouter Fr√®re", command=self.add_sibling).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Ajouter Enfant", command=self.add_child).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Importer Masse", command=self.import_bulk_children).pack(side=tk.LEFT, padx=5) # Nouveau bouton
        tk.Button(button_frame, text="Modifier", command=self.edit_node).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Supprimer", command=self.delete_node).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Sauvegarder", command=self.save_data).pack(side=tk.RIGHT, padx=5)

        # ====== Cadre du Fil d'Ariane (Breadcrumb) ======
        breadcrumb_outer_frame = tk.Frame(self.root, bg="#f0f0f0", relief=tk.GROOVE, borderwidth=1)
        breadcrumb_outer_frame.pack(fill=tk.X, padx=10, pady=(5, 2))
        
        tk.Label(breadcrumb_outer_frame, text="üìç Chemin:", bg="#f0f0f0", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(5, 2))
        
        # Cadre interne pour les liens cliquables du breadcrumb
        self.breadcrumb_frame = tk.Frame(breadcrumb_outer_frame, bg="#f0f0f0")
        self.breadcrumb_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=3)
        
        # Label par d√©faut quand rien n'est s√©lectionn√©
        self.breadcrumb_placeholder = tk.Label(self.breadcrumb_frame, text="S√©lectionnez un √©l√©ment pour voir son chemin", fg="gray", bg="#f0f0f0", font=("Segoe UI", 9, "italic"))
        self.breadcrumb_placeholder.pack(side=tk.LEFT)

        # Arbre (Treeview)
        # Colonnes: Part Number (sera dans la colonne #0 pour l'arborescence), Position, Description
        self.tree = ttk.Treeview(self.root, columns=("Position", "Description"))
        
        # Configuration de la colonne #0 (l'arbre lui-m√™me) -> Part Number
        self.tree.heading("#0", text="Part Number (Arbre)", anchor=tk.W)
        self.tree.heading("Position", text="Position", anchor=tk.W)
        self.tree.heading("Description", text="Description", anchor=tk.W)
        
        # Configuration des colonnes
        self.tree.column("#0", stretch=tk.YES, width=300) 
        self.tree.column("Position", stretch=tk.NO, width=80) # Position est souvent court
        self.tree.column("Description", stretch=tk.YES, width=420)

        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Ajout d'une barre de d√©filement verticale
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)  
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.place(relx=1.0, rely=0.0, relheight=1.0, anchor="ne") 

        # Dictionnaire pour stocker les donn√©es en m√©moire (id -> {parent_id, position, part_number, description})
        self.data_store = {}

        # Liaison de l'√©v√©nement de s√©lection pour mettre √† jour le breadcrumb
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        # Chargement initial des donn√©es
        self.load_data()

    def load_data(self):
        """Charge les donn√©es depuis le fichier Excel s'il existe."""
        if os.path.exists(EXCEL_FILE):
            try:
                df = pd.read_excel(EXCEL_FILE)
                # Assurons-nous que le fichier n'est pas vide et a les bonnes colonnes (ou compatibles)
                if not df.empty:
                    # Nettoyer l'arbre actuel
                    for item in self.tree.get_children():
                        self.tree.delete(item)
                    self.data_store = {}

                    # Convertir le DataFrame en dictionnaire pour un acc√®s rapide
                    nodes = {}
                    for index, row in df.iterrows():
                        node_id = str(row["ID"])
                        parent_id = str(row["ParentID"]) if "ParentID" in df.columns and pd.notna(row["ParentID"]) else ""
                        part_number = str(row["PartNumber"]) if "PartNumber" in df.columns else ""
                        description = str(row["Description"]) if "Description" in df.columns and pd.notna(row["Description"]) else ""
                        position = str(row["Position"]) if "Position" in df.columns and pd.notna(row["Position"]) else ""

                        nodes[node_id] = {"parent_id": parent_id, "part_number": part_number, "description": description, "position": position}

                    # Reconstruire l'arbre
                    to_add = list(nodes.keys())
                    added = set()

                    # On ajoute d'abord les racines (parent_id vide ou non trouv√© dans nodes)
                    for node_id in to_add[:]:
                        parent_id = nodes[node_id]["parent_id"]
                        if not parent_id or parent_id not in nodes:
                            self.insert_node_in_tree("", node_id, nodes[node_id]["position"], nodes[node_id]["part_number"], nodes[node_id]["description"])
                            self.data_store[node_id] = nodes[node_id]
                            added.add(node_id)
                            to_add.remove(node_id)
                    
                    # Ensuite on ajoute les enfants it√©rativement
                    last_count = len(to_add) + 1
                    while to_add:
                        current_count = len(to_add)
                        if current_count == last_count:
                            print("Attention: Des orphelins ont √©t√© d√©tect√©s et ignor√©s.")
                            break 
                        last_count = current_count

                        for node_id in to_add[:]:
                            parent_id = nodes[node_id]["parent_id"]
                            if parent_id in added:
                                self.insert_node_in_tree(parent_id, node_id, nodes[node_id]["position"], nodes[node_id]["part_number"], nodes[node_id]["description"])
                                self.data_store[node_id] = nodes[node_id]
                                added.add(node_id)
                                to_add.remove(node_id)

            except Exception as e:
                messagebox.showerror("Erreur de chargement", f"Impossible de charger le fichier Excel:\n{e}")

    # ====== M√©thodes pour le Fil d'Ariane (Breadcrumb) ======
    
    def on_tree_select(self, event=None):
        """Appel√© quand un √©l√©ment est s√©lectionn√© dans l'arbre."""
        selected_items = self.tree.selection()
        if selected_items:
            self.update_breadcrumb(selected_items[0])
        else:
            self.clear_breadcrumb()
    
    def get_ancestors(self, node_id):
        """Retourne la liste des anc√™tres du noeud (du plus ancien au plus r√©cent), incluant le noeud lui-m√™me."""
        path = []
        current_id = node_id
        
        while current_id:
            if current_id in self.data_store:
                path.insert(0, current_id)
                current_id = self.data_store[current_id].get("parent_id", "")
            else:
                # Noeud racine ou non trouv√© dans data_store
                if self.tree.exists(current_id):
                    path.insert(0, current_id)
                break
        
        return path
    
    def update_breadcrumb(self, node_id):
        """Met √† jour l'affichage du fil d'Ariane pour le noeud s√©lectionn√©.
        Affiche le chemin de l'√©l√©ment actuel vers la racine (ex: Pi√®ce ‚Üí Section ‚Üí Module ‚Üí TETRATOP 4)
        """
        # Nettoyer le cadre breadcrumb
        for widget in self.breadcrumb_frame.winfo_children():
            widget.destroy()
        
        # Obtenir le chemin complet (de la racine vers l'√©l√©ment)
        path = self.get_ancestors(node_id)
        
        if not path:
            self.breadcrumb_placeholder = tk.Label(
                self.breadcrumb_frame, 
                text="S√©lectionnez un √©l√©ment pour voir son chemin", 
                fg="gray", bg="#f0f0f0", font=("Segoe UI", 9, "italic")
            )
            self.breadcrumb_placeholder.pack(side=tk.LEFT)
            return
        
        # INVERSER le chemin : de l'√©l√©ment actuel vers la racine
        path = list(reversed(path))
        
        # Cr√©er les liens pour chaque √©l√©ment du chemin
        for i, ancestor_id in enumerate(path):
            data = self.data_store.get(ancestor_id, {})
            # Utiliser la DESCRIPTION au lieu du Part Number pour l'affichage
            description = data.get("description", "")
            part_number = data.get("part_number", self.tree.item(ancestor_id, "text"))
            # Si pas de description, utiliser le part number comme fallback
            display_text = description if description else part_number
            
            # Style diff√©rent pour le premier √©l√©ment (√©l√©ment s√©lectionn√©)
            is_current = (i == 0)
            
            if is_current:
                # √âl√©ment actuel : en gras, non cliquable
                label = tk.Label(
                    self.breadcrumb_frame, 
                    text=display_text,
                    bg="#e3f2fd",  # Bleu clair
                    fg="#1565c0",  # Bleu fonc√©
                    font=("Segoe UI", 9, "bold"),
                    padx=6, pady=2,
                    relief=tk.SOLID,
                    borderwidth=1
                )
            else:
                # Anc√™tre : cliquable avec effet hover
                label = tk.Label(
                    self.breadcrumb_frame, 
                    text=display_text,
                    bg="#f0f0f0",
                    fg="#1976d2",  # Bleu lien
                    font=("Segoe UI", 9, "underline"),
                    padx=4, pady=2,
                    cursor="hand2"
                )
                # Lier le clic pour naviguer vers cet anc√™tre
                label.bind("<Button-1>", lambda e, aid=ancestor_id: self.navigate_to_node(aid))
                # Effet hover
                label.bind("<Enter>", lambda e, lbl=label: lbl.config(bg="#e0e0e0", fg="#0d47a1"))
                label.bind("<Leave>", lambda e, lbl=label: lbl.config(bg="#f0f0f0", fg="#1976d2"))
            
            label.pack(side=tk.LEFT, padx=1)
            
            # Ajouter le s√©parateur " - " sauf pour le dernier √©l√©ment (la racine)
            if i < len(path) - 1:
                separator = tk.Label(
                    self.breadcrumb_frame, 
                    text="  -  ",
                    bg="#f0f0f0",
                    fg="#888888",
                    font=("Segoe UI", 9)
                )
                separator.pack(side=tk.LEFT)
    
    def clear_breadcrumb(self):
        """Efface le fil d'Ariane et affiche le placeholder."""
        for widget in self.breadcrumb_frame.winfo_children():
            widget.destroy()
        
        self.breadcrumb_placeholder = tk.Label(
            self.breadcrumb_frame, 
            text="S√©lectionnez un √©l√©ment pour voir son chemin", 
            fg="gray", bg="#f0f0f0", font=("Segoe UI", 9, "italic")
        )
        self.breadcrumb_placeholder.pack(side=tk.LEFT)
    
    def navigate_to_node(self, node_id):
        """Navigue vers un noeud sp√©cifique : le s√©lectionne et le rend visible."""
        if self.tree.exists(node_id):
            # Ouvrir tous les parents pour rendre le noeud visible
            parent_id = self.tree.parent(node_id)
            while parent_id:
                self.tree.item(parent_id, open=True)
                parent_id = self.tree.parent(parent_id)
            
            # S√©lectionner le noeud
            self.tree.selection_set(node_id)
            # Le rendre visible (scroll)
            self.tree.see(node_id)
            # Donner le focus √† l'arbre
            self.tree.focus(node_id)

    def get_path_string(self, node_id):
        """Retourne le chemin lisible d'un noeud (descriptions s√©par√©es par ' - ')."""
        path = self.get_ancestors(node_id)
        path_names = []
        for ancestor_id in path:
            data = self.data_store.get(ancestor_id, {})
            description = data.get("description", "")
            part_number = data.get("part_number", "")
            # Utiliser la description, sinon le part number
            name = description if description else part_number
            if name:
                path_names.append(name)
        # Inverser pour avoir : √âl√©ment - Parent - ... - Racine
        path_names.reverse()
        return " - ".join(path_names)

    def save_data(self):
        """Sauvegarde les donn√©es dans Excel avec groupement, chemin lisible et mise en forme."""
        export_data = []
        
        # Parcours r√©cursif pour obtenir l'ordre visuel et les niveaux
        def traverse(parent_id, level):
            children = self.tree.get_children(parent_id)
            for child_id in children:
                node_data = self.data_store[child_id]
                # G√©n√©rer le chemin lisible (stock√© pour les commentaires Excel)
                chemin = self.get_path_string(child_id)
                export_data.append({
                    "Position": node_data.get("position", ""),
                    "PartNumber": node_data["part_number"],
                    "Description": node_data["description"],
                    "Niveau": level,
                    "_Level": level,
                    "_ID": child_id,
                    "_ParentID": node_data["parent_id"],
                    "_Chemin": chemin  # Utilis√© pour les commentaires, pas affich√© en colonne
                })
                traverse(child_id, level + 1)

        traverse("", 0)

        if not export_data:
            # Si vide, on cr√©e juste les headers
            df = pd.DataFrame(columns=["Position", "PartNumber", "Description", "Niveau"])
            df.to_excel(EXCEL_FILE, index=False)
            messagebox.showinfo("Sauvegarde", "Fichier Excel sauvegard√© (vide).")
            return

        # Colonnes dans l'ordre souhait√© (sans colonne Chemin - sera en commentaire)
        df = pd.DataFrame(export_data, columns=["Position", "PartNumber", "Description", "Niveau", "_Level", "_ID", "_ParentID", "_Chemin"])
        
        try:
            # On sauvegarde sans les colonnes internes
            df_to_save = df.drop(columns=["_Level", "_ID", "_ParentID", "_Chemin"])
            
            # Post-traitement avec OpenPyXL pour le groupement et la mise en forme
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils.dataframe import dataframe_to_rows
                from openpyxl.comments import Comment
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Arborescence Pi√®ces"
                
                # === ZONE D'AFFICHAGE DU CHEMIN EN HAUT ===
                # Ligne 1 : Titre de la zone chemin
                ws.merge_cells('A1:D1')
                ws['A1'] = "üìç CHEMIN : Cliquez sur une ligne puis regardez le commentaire (triangle rouge) ‚Üí"
                ws['A1'].font = Font(bold=True, size=11, color="0052A3")
                ws['A1'].fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[1].height = 25
                
                # Ligne 2 : Vide (s√©paration)
                ws.row_dimensions[2].height = 5
                
                # √âcrire les en-t√™tes √† la ligne 3
                headers = ["Position", "PartNumber", "Description", "Niveau"]
                for c_idx, header in enumerate(headers, 1):
                    ws.cell(row=3, column=c_idx, value=header)
                
                # √âcrire les donn√©es √† partir de la ligne 4
                for r_idx, row_data in enumerate(export_data):
                    excel_row = r_idx + 4  # Donn√©es commencent √† la ligne 4
                    ws.cell(row=excel_row, column=1, value=row_data["Position"])
                    ws.cell(row=excel_row, column=2, value=row_data["PartNumber"])
                    ws.cell(row=excel_row, column=3, value=row_data["Description"])
                    ws.cell(row=excel_row, column=4, value=row_data["Niveau"])
                    
                    # Ajouter un COMMENTAIRE avec le chemin sur la colonne Description
                    chemin = row_data["_Chemin"]
                    comment = Comment(f"üìç CHEMIN:\n{chemin}", "Syst√®me")
                    comment.width = 400
                    comment.height = 80
                    ws.cell(row=excel_row, column=3).comment = comment
                
                # Style pour l'en-t√™te (ligne 3)
                header_fill = PatternFill(start_color="0052A3", end_color="0052A3", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                for cell in ws[3]:
                    if cell.value:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                
                # Style pour les donn√©es
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                
                # Couleurs altern√©es pour les niveaux
                level_colors = [
                    "FFFFFF",  # Niveau 0 - Blanc
                    "E3F2FD",  # Niveau 1 - Bleu tr√®s clair
                    "BBDEFB",  # Niveau 2 - Bleu clair
                    "90CAF9",  # Niveau 3 - Bleu
                    "64B5F6",  # Niveau 4 - Bleu moyen
                    "42A5F5",  # Niveau 5+ - Bleu plus fonc√©
                ]
                
                for i, row_data in enumerate(export_data):
                    row_idx = i + 4  # Donn√©es commencent √† la ligne 4
                    level = row_data["_Level"]
                    color_idx = min(level, len(level_colors) - 1)
                    fill = PatternFill(start_color=level_colors[color_idx], end_color=level_colors[color_idx], fill_type="solid")
                    
                    for cell in ws[row_idx]:
                        cell.border = thin_border
                        cell.fill = fill
                        cell.alignment = Alignment(vertical="center")
                
                # Ajuster la largeur des colonnes
                column_widths = {
                    'A': 12,   # Position
                    'B': 20,   # PartNumber
                    'C': 45,   # Description (plus large car c'est l√† qu'on clique)
                    'D': 10,   # Niveau
                }
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                
                # Figer les lignes 1-3 (zone chemin + en-t√™tes)
                ws.freeze_panes = 'A4'
                
                # Activer les filtres automatiques sur les en-t√™tes (ligne 3)
                ws.auto_filter.ref = f"A3:D{3 + len(export_data)}"
                
                # IMPORTANT: Pour que les '+' soient sur la ligne du parent (en haut du groupe)
                ws.sheet_properties.outlinePr.summaryBelow = False
                
                # Appliquer les niveaux de plan (grouping) - donn√©es commencent ligne 4
                for i, row_data in enumerate(export_data):
                    level = row_data["_Level"]
                    if level > 0:
                        ws.row_dimensions[i + 4].outlineLevel = level
                
                wb.save(EXCEL_FILE)
                messagebox.showinfo("Sauvegarde", f"Fichier Excel sauvegard√© avec succ√®s!\n\n‚Ä¢ {len(export_data)} √©l√©ments\n‚Ä¢ Survolez la colonne Description pour voir le chemin\n‚Ä¢ Triangle rouge = commentaire avec chemin\n‚Ä¢ Groupement par niveau activ√©")
                
            except ImportError:
                # Fallback si openpyxl n'est pas disponible avec les styles
                df_to_save.to_excel(EXCEL_FILE, index=False)
                messagebox.showinfo("Sauvegarde", f"Fichier Excel sauvegard√© (simple).\n{len(export_data)} √©l√©ments.")
                
        except Exception as e:
            messagebox.showerror("Erreur de sauvegarde", f"Impossible de sauvegarder le fichier Excel:\n{e}\nV√©rifiez que le fichier n'est pas ouvert ailleurs.")

    def insert_node_in_tree(self, parent_id, node_id, position, part_number, description):
        """Helper pour ins√©rer dans l'arbre visuel."""
        # text=part_number pour l'afficher dans la colonne #0 (l'arbre)
        # values=(position, description)
        self.tree.insert(parent_id, 'end', iid=node_id, text=part_number, values=(position, description))
        
        # On ouvre le parent pour montrer le nouvel enfant
        if parent_id:
            self.tree.item(parent_id, open=True)

    def add_root(self):
        self.prompt_and_add_node("")

    def add_sibling(self):
        """Ajoute un noeud au m√™me niveau que la s√©lection actuelle."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("S√©lection requise", "Veuillez s√©lectionner un noeud pour lui ajouter un fr√®re.")
            return
        
        node_id = selected_item[0]
        # Trouver le parent du noeud s√©lectionn√©
        parent_id = self.tree.parent(node_id)
        
        self.prompt_and_add_node(parent_id)

    def add_child(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("S√©lection requise", "Veuillez s√©lectionner un noeud parent.")
            return
        parent_id = selected_item[0]
        self.prompt_and_add_node(parent_id)

    def import_bulk_children(self):
        """Ouvre une fen√™tre pour importer plusieurs enfants d'un coup via copier-coller."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("S√©lection requise", "Veuillez s√©lectionner un noeud parent pour les items import√©s.")
            return
        parent_id = selected_item[0]
        
        # Fen√™tre de dialogue
        dialog = tk.Toplevel(self.root)
        dialog.title("Importer en masse via texte")
        dialog.geometry("700x500")

        tk.Label(dialog, text="Collez vos donn√©es ci-dessous :").pack(anchor=tk.W, padx=10, pady=(10, 0))
        text_area = tk.Text(dialog, height=15)
        text_area.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Cadre pour les param√®tres de d√©coupage
        frame = tk.LabelFrame(dialog, text="D√©limitation des colonnes (index de caract√®re, 0 = d√©but de ligne)")
        frame.pack(fill=tk.X, padx=10, pady=10)
        
        grid_frame = tk.Frame(frame)
        grid_frame.pack(padx=5, pady=5)

        # Ligne 1: Position
        tk.Label(grid_frame, text="Index D√©but Position :").grid(row=0, column=0, sticky=tk.E, padx=5)
        entry_pos_start = tk.Entry(grid_frame, width=5)
        entry_pos_start.insert(0, "0") 
        entry_pos_start.grid(row=0, column=1, padx=5)

        tk.Label(grid_frame, text="Index Fin Position :").grid(row=0, column=2, sticky=tk.E, padx=5)
        entry_pos_end = tk.Entry(grid_frame, width=5)
        entry_pos_end.insert(0, "3") 
        entry_pos_end.grid(row=0, column=3, padx=5)

        # Ligne 2: Part Number
        tk.Label(grid_frame, text="Index D√©but Part Number :").grid(row=1, column=0, sticky=tk.E, padx=5)
        entry_pn_start = tk.Entry(grid_frame, width=5)
        entry_pn_start.insert(0, "4") 
        entry_pn_start.grid(row=1, column=1, padx=5)

        tk.Label(grid_frame, text="Index Fin Part Number :").grid(row=1, column=2, sticky=tk.E, padx=5)
        entry_pn_end = tk.Entry(grid_frame, width=5)
        entry_pn_end.insert(0, "16") 
        entry_pn_end.grid(row=1, column=3, padx=5)

        # Ligne 3: Description
        tk.Label(grid_frame, text="Index D√©but Description :").grid(row=2, column=0, sticky=tk.E, padx=5)
        entry_desc_start = tk.Entry(grid_frame, width=5)
        entry_desc_start.insert(0, "19") 
        entry_desc_start.grid(row=2, column=1, padx=5)

        tk.Label(grid_frame, text="Caract√®re de fin (Optionnel) :").grid(row=2, column=2, sticky=tk.E, padx=5)
        entry_desc_end_char = tk.Entry(grid_frame, width=5)
        entry_desc_end_char.insert(0, ".") 
        entry_desc_end_char.grid(row=2, column=3, padx=5)

        # Configuration des tags pour les couleurs
        text_area.tag_configure("pos", background="#ffeb3b") # Jaune
        text_area.tag_configure("pn", background="#80deea")  # Cyan
        text_area.tag_configure("desc", background="#c8e6c9") # Vert clair

        def update_highlights(event=None):
            text_area.tag_remove("pos", "1.0", tk.END)
            text_area.tag_remove("pn", "1.0", tk.END)
            text_area.tag_remove("desc", "1.0", tk.END)
            
            try:
                pos_start = int(entry_pos_start.get())
                pos_end = int(entry_pos_end.get())
                pn_start = int(entry_pn_start.get())
                pn_end = int(entry_pn_end.get())
                desc_start = int(entry_desc_start.get())
                stop_char = entry_desc_end_char.get()
            except ValueError:
                return 

            content = text_area.get("1.0", tk.END)
            lines = content.split('\n')
            
            for i, line in enumerate(lines):
                line_idx = i + 1
                if not line.strip(): continue
                
                # Position tagging
                if len(line) > pos_start:
                    p_end = min(len(line), pos_end)
                    if p_end > pos_start:
                        text_area.tag_add("pos", f"{line_idx}.{pos_start}", f"{line_idx}.{p_end}")
                
                # PN tagging
                if len(line) > pn_start:
                    p_end = min(len(line), pn_end)
                    if p_end > pn_start:
                        text_area.tag_add("pn", f"{line_idx}.{pn_start}", f"{line_idx}.{p_end}")
                
                # Desc tagging
                if len(line) > desc_start:
                    raw_desc = line[desc_start:]
                    d_end = len(line)
                    
                    if stop_char:
                        found_idx = raw_desc.find(stop_char)
                        if found_idx != -1:
                            d_end = desc_start + found_idx
                    
                    if d_end > desc_start:
                        text_area.tag_add("desc", f"{line_idx}.{desc_start}", f"{line_idx}.{d_end}")

        # Bindings pour mise √† jour temps r√©el
        text_area.bind('<KeyRelease>', update_highlights)
        entry_pos_start.bind('<KeyRelease>', update_highlights)
        entry_pos_end.bind('<KeyRelease>', update_highlights)
        entry_pn_start.bind('<KeyRelease>', update_highlights)
        entry_pn_end.bind('<KeyRelease>', update_highlights)
        entry_desc_start.bind('<KeyRelease>', update_highlights)
        entry_desc_end_char.bind('<KeyRelease>', update_highlights)

        # Appel initial pour colorer si des valeurs par d√©faut sont pr√©sentes
        dialog.after(100, update_highlights) # Petite pause pour s'assurer que la fen√™tre est rendue

        def do_import():
            try:
                raw_text = text_area.get("1.0", tk.END).strip()
                if not raw_text: return
                
                pos_start = int(entry_pos_start.get())
                pos_end = int(entry_pos_end.get())
                pn_start = int(entry_pn_start.get())
                pn_end = int(entry_pn_end.get())
                desc_start = int(entry_desc_start.get())
                stop_char = entry_desc_end_char.get()
                
                lines = raw_text.split('\n')
                count = 0
                for line in lines:
                    line = line.rstrip() 
                    if not line: continue
                    try:
                        # 1. Extraction Position
                        position = ""
                        if len(line) > pos_start:
                            current_pos_end = min(len(line), pos_end)
                            position = line[pos_start:current_pos_end].strip()

                        # 2. Extraction Part Number - Condition critique
                        if len(line) <= pn_start: continue 
                        
                        current_pn_end = min(len(line), pn_end)
                        part_number = line[pn_start:current_pn_end].strip()
                        
                        # 3. Extraction Description
                        description = ""
                        if len(line) > desc_start:
                            raw_desc = line[desc_start:]
                            if stop_char and stop_char in raw_desc:
                                description = raw_desc.split(stop_char, 1)[0].strip()
                            else:
                                description = raw_desc.strip()
                        
                        if part_number: 
                            new_id = str(uuid.uuid4())
                            # Update parameters including position
                            self.insert_node_in_tree(parent_id, new_id, position, part_number, description)
                            self.data_store[new_id] = {
                                "parent_id": parent_id,
                                "position": position,
                                "part_number": part_number,
                                "description": description
                            }
                            count += 1
                    except IndexError:
                        continue 
                
                self.save_data() 
                messagebox.showinfo("Succ√®s", f"{count} √©l√©ments import√©s.")
                dialog.destroy()
            except ValueError:
                messagebox.showerror("Erreur", "Veuillez entrer des nombres entiers valides pour les index.")

        tk.Button(dialog, text="Lancer l'Importation", command=do_import, bg="#dddddd", height=2).pack(pady=10)

    def prompt_and_add_node(self, parent_id):
        # Dialogue pour Position
        position = simpledialog.askstring("Nouveau Noeud", "Entrez la Position (Optionnel):", initialvalue="")
        if position is None: position = ""

        # Dialogue pour Part Number
        part_number = simpledialog.askstring("Nouveau Noeud", "Entrez le Part Number:")
        if part_number is None: return # Annul√©

        # Dialogue pour Description
        description = simpledialog.askstring("Nouveau Noeud", "Entrez la Description:")
        if description is None: description = ""

        # Cr√©ation
        new_id = str(uuid.uuid4())
        self.insert_node_in_tree(parent_id, new_id, position, part_number, description)
        
        # Mise √† jour des donn√©es
        self.data_store[new_id] = {
            "parent_id": parent_id,
            "position": position,
            "part_number": part_number,
            "description": description
        }

        # Sauvegarde temps r√©el
        self.save_data()

    def edit_node(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("S√©lection requise", "Veuillez s√©lectionner un noeud √† modifier.")
            return
        node_id = selected_item[0]
        
        # R√©cup√©rer les donn√©es
        current_data = self.data_store.get(node_id)
        if not current_data: return 

        new_pos = simpledialog.askstring("Modifier", "Modifier Position:", initialvalue=current_data.get("position", ""))
        if new_pos is None: return

        new_pn = simpledialog.askstring("Modifier", "Modifier Part Number:", initialvalue=current_data["part_number"])
        if new_pn is None: return

        new_desc = simpledialog.askstring("Modifier", "Modifier Description:", initialvalue=current_data["description"])
        if new_desc is None: return

        # Mise √† jour arbre
        self.tree.item(node_id, text=new_pn, values=(new_pos, new_desc))

        # Mise √† jour donn√©es
        self.data_store[node_id]["position"] = new_pos
        self.data_store[node_id]["part_number"] = new_pn
        self.data_store[node_id]["description"] = new_desc

        # Sauvegarde
        self.save_data()

    def delete_node(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("S√©lection requise", "Veuillez s√©lectionner un noeud √† supprimer.")
            return
        node_id = selected_item[0]
        
        if messagebox.askyesno("Confirmation", "Voulez-vous vraiment supprimer ce noeud et TOUS ses enfants?"):
            self.delete_node_recursive(node_id)
            self.save_data()

    def delete_node_recursive(self, node_id):
        # Supprimer visuellement (supprime aussi les enfants dans Treeview)
        # Mais on doit aussi nettoyer self.data_store r√©cursivement
        
        # Trouver les enfants dans data_store
        children = [k for k, v in self.data_store.items() if v["parent_id"] == node_id]
        for child in children:
            self.delete_node_recursive(child)
            
        # Supprimer soi-m√™me
        if node_id in self.data_store:
            del self.data_store[node_id]
        
        # Si le noeud existe encore dans l'arbre (c'est la racine de la suppression), on le supprime
        if self.tree.exists(node_id):
            self.tree.delete(node_id)

if __name__ == "__main__":
    # V√©rification des d√©pendances au lancement
    try:
        import pandas
        import openpyxl
    except ImportError:
        messagebox.showerror("Erreur de d√©pendances", "Les biblioth√®ques 'pandas' et 'openpyxl' sont requises.\nVeuillez les installer avec: pip install pandas openpyxl")
        exit()

    root = tk.Tk()
    app = TreeApp(root)
    root.mainloop()
